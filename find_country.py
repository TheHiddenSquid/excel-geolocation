import sys
import geopy
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
import requests


def translate_country_name(country_name):
    if country_name == None:
        return "None"
    response = requests.get(f"https://translate.googleapis.com/translate_a/single?client=gtx&sl=auto&tl=sv&dt=t&q={country_name}")
    translation = response.json()[0][0][0]
    return translation

def list_from_excell_col(excell_sheet:str, colls:list[str]):
    result = []   
    wb = load_workbook(excell_sheet)
    ws = wb.active
    for col in colls:
        excellcol = ws[col]
        column_list = [excellcol[x].value for x in range(len(excellcol))]
        adresses = column_list[1:3738]
        result.append(adresses)

    return result

def find_country(loc):
    geolocator = Nominatim(user_agent="me")

    if loc == None:
        return None
    try:
        location = geolocator.geocode(loc)
        fulladdress = location.address
        country = fulladdress.split(',')[-1]
        return country
    except KeyboardInterrupt:
        sys.exit()
    except:
        return None
        
def main():
    col1, col2, col3 = list_from_excell_col("country_test.xlsx", ["B", "C", "D"])
    num_countries = len(col1)

    for i in range(num_countries):
        ans = ""
        cases = [find_country(col1[i]), find_country(col2[i]), find_country(col3[i])]
        if cases[0]==cases[1] and cases[1]==cases[2]:
            ans = cases[0]
        else:
            nones = cases.count(None)
            for _ in range(nones):
                cases.remove(None)
            if len(cases)<2: # set to ==0 for more hits
                ans = None
            elif len(cases) == cases.count(cases[0]):
                ans =  cases[0]
            else:
                 ans = None

        with open("results.txt", "a") as f:
            f.write(f"{translate_country_name(ans)}\n")

if __name__ == "__main__":
    main()
